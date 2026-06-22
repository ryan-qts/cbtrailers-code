"""
Step 3: Apply reviewed image choices back to the Excel file.

Usage: python apply_images.py
"""

import openpyxl
import json
import os

INPUT_FILE    = r"C:\Users\jerem\PyCharmMiscProject\scrape\imageScraping2\MasterList_import_ready_.xlsx"
REVIEWED_FILE = r"C:\Users\jerem\PyCharmMiscProject\scrape\imageScraping2\image_reviewed.json"

def main():
    if not os.path.exists(REVIEWED_FILE):
        print("No reviewed file found. Complete the review tool first.")
        return

    with open(REVIEWED_FILE, encoding='utf-8') as f:
        reviewed = json.load(f)

    print(f"Loading {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    img_idx = headers.index('image') + 1

    applied = skipped = 0
    for key, data in reviewed.items():
        if not data.get('image'):
            skipped += 1
            continue
        row_num = int(key)
        ws.cell(row=row_num, column=img_idx).value = data['image']
        applied += 1

    wb.save(INPUT_FILE)
    print(f"Done! Applied {applied} images, skipped {skipped} (no selection).")
    print(f"Saved: {INPUT_FILE}")

if __name__ == "__main__":
    main()
