"""
VisionLED Scraper — Playwright version
---------------------------------------
Install:
    pip install playwright beautifulsoup4 pandas openpyxl lxml
    playwright install chromium

Usage:
    Place visionled-sheet.xlsx in the same folder, then:
    python visionled_scraper.py
"""

import re
import asyncio
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

# ─── CONFIG ───────────────────────────────────────────────────────────────────
INPUT_FILE    = "visionled-sheet.xlsx"
OUTPUT_FILE   = "visionled_updated.xlsx"
PRODUCTS_URL  = "https://visionled-lights.com/products/"
SCROLL_PAUSE  = 2500   # ms to wait after each scroll
SCROLL_STABLE = 5      # stop after this many scrolls with zero new links
PAGE_DELAY    = 1200   # ms between product page loads
HEADLESS      = False  # visible browser passes captcha better


# ─── NORMALISER ───────────────────────────────────────────────────────────────
def clean_sku(sku):
    if not isinstance(sku, str):
        return ""
    return re.sub(r"\s+", "", sku).upper().strip()


# ─── LOAD EXCEL ───────────────────────────────────────────────────────────────
print("Loading Excel file …")
df = pd.read_excel(INPUT_FILE, dtype=str)
df.columns = df.columns.str.strip()

if "REF" not in df.columns:
    raise ValueError(f"'REF' column not found. Columns: {list(df.columns)}")

for col in ["Image URL", "Box Qty", "Inner Box Qty", "Product URL"]:
    if col not in df.columns:
        df[col] = ""

sku_index: dict[str, list[int]] = {}
for idx, raw in enumerate(df["REF"]):
    key = clean_sku(str(raw))
    if key:
        sku_index.setdefault(key, []).append(idx)

print(f"  {len(df)} rows | {len(sku_index)} unique SKUs\n")


# ─── BROWSER HELPERS ─────────────────────────────────────────────────────────
async def safe_goto(page, url):
    """Navigate, retry up to 4x until page is not a captcha redirect."""
    for attempt in range(4):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        except PWTimeout:
            print(f"  Timeout (attempt {attempt+1})")
            continue
        await page.wait_for_timeout(1500)
        content = await page.content()
        if "sgcaptcha" not in content:
            return content
        print(f"  Captcha (attempt {attempt+1}) — waiting …")
        await page.wait_for_timeout(5000)
    return await page.content()


def extract_links_from_html(html):
    """Pull all hrefs that look like individual product pages."""
    soup = BeautifulSoup(html, "lxml")
    links = set()
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if (
            "/products/" in href
            and href.rstrip("/") != PRODUCTS_URL.rstrip("/")
            and not href.endswith("/products/")
            and "?" not in href
            and "#" not in href
        ):
            if not href.startswith("http"):
                href = "https://visionled-lights.com" + href
            links.add(href)
    return links


async def collect_all_product_links(page):
    """
    Scroll the /products/ listing until no new links appear for
    SCROLL_STABLE consecutive scrolls, then return all found URLs.
    """
    print(f"Opening listing page: {PRODUCTS_URL}")
    await safe_goto(page, PRODUCTS_URL)

    all_links = set()
    stable = 0
    scroll_n = 0

    while stable < SCROLL_STABLE:
        scroll_n += 1

        # Step-scroll (better than jumping to bottom for lazy-load triggers)
        await page.evaluate("window.scrollBy(0, window.innerHeight * 3)")
        await page.wait_for_timeout(SCROLL_PAUSE)

        # Every 3 scrolls also hard-jump to the very bottom
        if scroll_n % 3 == 0:
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(SCROLL_PAUSE)

        new_links = extract_links_from_html(await page.content())
        added = new_links - all_links

        if added:
            all_links |= added
            stable = 0
            print(f"  Scroll {scroll_n}: +{len(added)} new  ({len(all_links)} total)")
        else:
            stable += 1
            print(f"  Scroll {scroll_n}: nothing new  (stable {stable}/{SCROLL_STABLE})")

    print(f"\nScroll complete — {len(all_links)} product pages found\n")
    return sorted(all_links)


# ─── PRODUCT PAGE PARSER ──────────────────────────────────────────────────────
def get_image(soup):
    for sel in [
        ".woocommerce-product-gallery__image img",
        ".woocommerce-product-gallery img",
        "img.wp-post-image",
        ".product img",
    ]:
        tag = soup.select_one(sel)
        if tag:
            url = tag.get("data-large_image") or tag.get("data-src") or tag.get("src")
            if url and "placeholder" not in url:
                return url
    og = soup.find("meta", property="og:image")
    return og["content"] if og else None


def parse_product_page(html):
    """
    Returns list of dicts: {sku, image_url, box_qty, inner_box_qty}
    Tries variation table first, then falls back to .sku element.
    """
    soup = BeautifulSoup(html, "lxml")
    image_url = get_image(soup)
    items = []

    # Primary: variation/spec table rows that start with a part-number pattern
    for table in soup.select("table"):
        for row in table.select("tr"):
            cols = [td.get_text(strip=True) for td in row.select("td")]
            if len(cols) < 2:
                continue
            if not re.match(r"^[A-Za-z]{1,8}[-]?\d", cols[0]):
                continue
            items.append({
                "sku":           clean_sku(cols[0]),
                "image_url":     image_url,
                "box_qty":       cols[2] if len(cols) > 2 else "",
                "inner_box_qty": cols[3] if len(cols) > 3 else "",
            })

    # Fallback: single .sku element on page
    if not items:
        sku_el = soup.select_one(".sku") or soup.select_one("[class*='sku']")
        if sku_el:
            raw = re.sub(r"(?i)^sku\s*[:\-]\s*", "", sku_el.get_text(strip=True))
            if raw:
                items.append({
                    "sku":           clean_sku(raw),
                    "image_url":     image_url,
                    "box_qty":       "",
                    "inner_box_qty": "",
                })

    return items


# ─── MAIN ─────────────────────────────────────────────────────────────────────
async def main():
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=HEADLESS,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
        )
        page = await context.new_page()

        # ── 1. Collect all product URLs ───────────────────────────────────────
        product_links = await collect_all_product_links(page)

        if not product_links:
            print("No products found — check the browser for a captcha.")
            await page.wait_for_timeout(10000)
            await browser.close()
            return

        # ── 2. Scrape each product page ───────────────────────────────────────
        matched_total  = 0
        unparsed_links = []   # pages where we got no items at all
        unmatched_skus = []   # pages parsed but SKU not in sheet (url, sku)

        for i, url in enumerate(product_links, 1):
            print(f"[{i}/{len(product_links)}] {url}")
            html  = await safe_goto(page, url)
            items = parse_product_page(html)

            if not items:
                print("  → could not parse any items")
                unparsed_links.append(url)
                await page.wait_for_timeout(PAGE_DELAY)
                continue

            for item in items:
                indices = sku_index.get(item["sku"], [])
                if indices:
                    for idx in indices:
                        df.at[idx, "Image URL"]     = item["image_url"] or ""
                        df.at[idx, "Box Qty"]       = item["box_qty"]
                        df.at[idx, "Inner Box Qty"] = item["inner_box_qty"]
                        df.at[idx, "Product URL"]   = url
                    matched_total += len(indices)
                    print(f"  ✓ {item['sku']} → {len(indices)} row(s)")
                else:
                    print(f"  ✗ {item['sku']} (not in sheet)")
                    unmatched_skus.append((url, item["sku"]))

            await page.wait_for_timeout(PAGE_DELAY)

        await browser.close()

    # ── 3. Save output ────────────────────────────────────────────────────────
    print(f"\nTotal rows updated: {matched_total}")
    df.to_excel(OUTPUT_FILE, index=False)

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    hfill = PatternFill("solid", fgColor="1F3864")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 60)
    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    print(f"Saved → {OUTPUT_FILE}")

    # ── 4. Summary report ─────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"  Product pages found : {len(product_links)}")
    print(f"  Rows updated        : {matched_total}")
    print(f"  Could not parse     : {len(unparsed_links)}")
    print(f"  Parsed but no match : {len(unmatched_skus)}")

    if unparsed_links:
        print("\n── Pages where NO items could be parsed ──")
        for link in unparsed_links:
            print(f"  {link}")

    if unmatched_skus:
        print("\n── Pages parsed but SKU not found in sheet ──")
        for link, sku in unmatched_skus:
            print(f"  {sku:30s}  {link}")

    print("=" * 60)


asyncio.run(main())
