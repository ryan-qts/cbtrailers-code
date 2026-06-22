"""
VisionLED Scraper — Playwright version
---------------------------------------
Uses a real Chromium browser to bypass the bot-detection captcha.

Install:
    pip install playwright beautifulsoup4 pandas openpyxl lxml
    playwright install chromium

Usage:
    Place visionled-sheet.xlsx in the same folder, then run:
    python visionled_scraper.py
"""

import re
import asyncio
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

# ─── CONFIG ───────────────────────────────────────────────────────────────────
INPUT_FILE   = "visionled-sheet.xlsx"
OUTPUT_FILE  = "visionled_updated.xlsx"
PRODUCTS_URL = "https://visionled-lights.com/products/"
PAGE_DELAY   = 1500   # ms between page loads
HEADLESS     = False  # keep False — visible browser is better at passing captcha

# ─── SKU NORMALISER ───────────────────────────────────────────────────────────
def clean_sku(sku):
    if not isinstance(sku, str):
        return ""
    return re.sub(r"\s+", "", sku).upper().strip()


# ─── LOAD EXCEL ───────────────────────────────────────────────────────────────
print("Loading Excel file …")
df = pd.read_excel(INPUT_FILE, dtype=str)
df.columns = df.columns.str.strip()

if "REF" not in df.columns:
    raise ValueError(f"'REF' column not found. Columns: {list(df.columns)}")

for col in ["Image URL", "Box Qty", "Inner Box Qty", "Product URL"]:
    if col not in df.columns:
        df[col] = ""

sku_index: dict[str, list[int]] = {}
for idx, raw in enumerate(df["REF"]):
    key = clean_sku(str(raw))
    if key:
        sku_index.setdefault(key, []).append(idx)

print(f"  {len(df)} rows, {len(sku_index)} unique SKUs\n")


# ─── HELPERS ──────────────────────────────────────────────────────────────────
async def safe_goto(page, url):
    """Navigate and wait until the page is not a captcha redirect."""
    for attempt in range(4):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        except PWTimeout:
            print(f"  Timeout on goto (attempt {attempt+1})")
            continue
        await page.wait_for_timeout(2000)
        content = await page.content()
        if "sgcaptcha" not in content:
            return content
        print(f"  Captcha detected (attempt {attempt+1}) — waiting …")
        await page.wait_for_timeout(4000)
    return await page.content()


def extract_product_links(html):
    soup = BeautifulSoup(html, "lxml")
    links = set()
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if (
            "/products/" in href
            and href.rstrip("/") != PRODUCTS_URL.rstrip("/")
            and not href.endswith("/products/")
            and "?" not in href
            and "#" not in href
        ):
            if not href.startswith("http"):
                href = "https://visionled-lights.com" + href
            links.add(href)
    return links


def parse_product_page(html, url):
    soup = BeautifulSoup(html, "lxml")
    items = []

    # ── Image ──
    img_tag = (
        soup.select_one(".woocommerce-product-gallery__image img") or
        soup.select_one(".woocommerce-product-gallery img") or
        soup.select_one("img.wp-post-image") or
        soup.select_one("figure img") or
        soup.select_one(".product img")
    )
    image_url = None
    if img_tag:
        image_url = img_tag.get("data-large_image") or img_tag.get("src")
    if not image_url:
        og = soup.find("meta", property="og:image")
        image_url = og["content"] if og else None

    # ── Variation table ──
    for table in soup.select("table"):
        for row in table.select("tr"):
            cols = [td.get_text(strip=True) for td in row.select("td")]
            if len(cols) < 2:
                continue
            if not re.match(r"^[A-Za-z]{1,8}[-]?\d", cols[0]):
                continue
            items.append({
                "sku":           clean_sku(cols[0]),
                "image_url":     image_url,
                "box_qty":       cols[2] if len(cols) > 2 else "",
                "inner_box_qty": cols[3] if len(cols) > 3 else "",
            })

    # ── Fallback: single-SKU product ──
    if not items:
        sku_el = soup.select_one(".sku") or soup.select_one("[class*='sku']")
        if sku_el:
            raw = re.sub(r"(?i)^sku\s*[:\-]\s*", "", sku_el.get_text(strip=True))
            if raw:
                items.append({
                    "sku":           clean_sku(raw),
                    "image_url":     image_url,
                    "box_qty":       "",
                    "inner_box_qty": "",
                })

    return items


# ─── MAIN ─────────────────────────────────────────────────────────────────────
async def main():
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=HEADLESS,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
        )
        page = await context.new_page()

        # ── Collect product URLs ──
        all_links = set()
        current_url = PRODUCTS_URL
        page_num = 1

        while current_url:
            print(f"Listing page {page_num}: {current_url}")
            html = await safe_goto(page, current_url)

            # Scroll to trigger lazy-load
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(PAGE_DELAY)
            html = await page.content()

            new_links = extract_product_links(html)
            added = new_links - all_links
            all_links |= new_links
            print(f"  +{len(added)} new links ({len(all_links)} total)")

            # Pagination
            soup = BeautifulSoup(html, "lxml")
            next_a = soup.select_one("a.next, a.page-numbers.next, a[rel='next']")
            if next_a and next_a.get("href"):
                next_href = next_a["href"]
                if next_href.rstrip("/") != current_url.rstrip("/"):
                    current_url = next_href
                    page_num += 1
                    continue
            break

        product_links = sorted(all_links)
        print(f"\nTotal product pages found: {len(product_links)}\n")

        if not product_links:
            print("No products found.")
            print("Tip: check the browser window — there may be a captcha to solve manually.")
            await page.wait_for_timeout(10000)
            await browser.close()
            return

        # ── Scrape each product ──
        matched_total = 0
        for i, url in enumerate(product_links, 1):
            print(f"[{i}/{len(product_links)}] {url}")
            html = await safe_goto(page, url)
            items = parse_product_page(html, url)

            if not items:
                print("  → no items parsed")
            else:
                for item in items:
                    indices = sku_index.get(item["sku"], [])
                    if indices:
                        for idx in indices:
                            df.at[idx, "Image URL"]     = item["image_url"] or ""
                            df.at[idx, "Box Qty"]       = item["box_qty"]
                            df.at[idx, "Inner Box Qty"] = item["inner_box_qty"]
                            df.at[idx, "Product URL"]   = url
                        matched_total += len(indices)
                        print(f"  ✓ {item['sku']} → {len(indices)} row(s)")
                    else:
                        print(f"  ✗ {item['sku']} (not in sheet)")

            await page.wait_for_timeout(PAGE_DELAY)

        await browser.close()

    # ── Save ──
    print(f"\nTotal rows updated: {matched_total}")
    df.to_excel(OUTPUT_FILE, index=False)

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    hfill = PatternFill("solid", fgColor="1F3864")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 60)
    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    print(f"Done → {OUTPUT_FILE}")


asyncio.run(main())