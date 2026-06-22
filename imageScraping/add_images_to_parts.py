"""
add_images_to_parts.py
----------------------
Searches for product images for each row in master_parts.xlsx using the
Parallel AI search API to find product pages, then fetches og:image meta
tags from those pages (which work even on JS-heavy sites).

REQUIREMENTS:
    pip install openpyxl requests pandas beautifulsoup4 lxml

USAGE:
    1. Place this script in the same folder as master_parts.xlsx
    2. Delete image_cache.json if it exists (fresh start)
    3. Run:  python add_images_to_parts.py
    4. Output: master_parts_with_images.xlsx

Progress is saved to image_cache.json every 25 rows — safe to interrupt and resume.
"""

import json
import os
import re
import time
import requests
from urllib.parse import urljoin
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Configuration ─────────────────────────────────────────────────────────────
PARALLEL_API_KEY = "J1xQ1_7gGzbBSIlJQ1c1Z4c8BdvStjEdeFSE0sZ_"
INPUT_FILE       = "master_parts.xlsx"
OUTPUT_FILE      = "master_parts_with_images.xlsx"
CACHE_FILE       = "image_cache.json"
SEARCH_DELAY     = 0.3    # seconds between Parallel API calls
FETCH_DELAY      = 0.2    # seconds between page fetches
MAX_PAGES        = 3      # max product pages to try per part
DEBUG            = False  # set True to see detailed output per row
# ──────────────────────────────────────────────────────────────────────────────

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
})

# Domains unlikely to have product images or that will block scraping
SKIP_DOMAINS = [
    "google.com", "facebook.com", "twitter.com", "youtube.com",
    "linkedin.com", "pinterest.com", "instagram.com", "reddit.com",
    "yelp.com", "bbb.org", "yellowpages.com", "wikipedia.org",
    "indeed.com", "glassdoor.com", "dnb.com", "manta.com",
]

# Fragment keywords that indicate a bad/non-product image
BAD_IMAGE_WORDS = [
    "logo", "icon", "favicon", "pixel", "tracking", "banner", "badge",
    "sprite", "button", "arrow", "spacer", "placeholder", "avatar",
    "no-image", "noimage", "missing", "default", "loading", "spinner",
    "background", "header", "footer", "nav", "menu", "star", "rating",
]


def load_cache() -> dict:
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE) as f:
            return json.load(f)
    return {}


def save_cache(cache: dict):
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2)


def is_good_image_url(url: str) -> bool:
    """True if the URL looks like a real product image."""
    if not url or not url.startswith("http"):
        return False
    low = url.lower()
    if any(bad in low for bad in BAD_IMAGE_WORDS):
        return False
    # Accept URLs with image extensions or CDN-style paths
    has_ext = bool(re.search(r'\.(jpg|jpeg|png|webp)(\?|$)', low))
    has_cdn = any(kw in low for kw in [
        "/images/", "/media/", "/product", "/catalog", "/photos/",
        "cdn.", "static.", "assets.", "img.", "imageservice", "cloudinary",
        "shopify", "bigcommerce", "woocommerce",
    ])
    return has_ext or has_cdn


def find_product_urls(vendor: str, part: str, desc: str) -> list:
    """Use Parallel to find candidate product page URLs."""
    v, p, d = vendor.strip(), part.strip(), desc.strip()

    queries = []
    if v and p:
        queries.append(f"{v} {p} product")
    if p and d:
        queries.append(f"{p} {d[:50]} buy")
    if d:
        queries.append(f"{d[:70]} trailer part")
    queries = list(dict.fromkeys(q for q in queries if q.strip()))[:4]

    if not queries:
        return []

    payload = {
        "objective": (
            f"Find product pages selling this trailer/industrial part: "
            f"Vendor={v or '?'}, Part={p or '?'}, Desc={d[:60] or '?'}. "
            f"Return URLs of e-commerce or distributor product pages."
        ),
        "search_queries": queries,
        "max_results": 10,
        "max_chars_per_result": 300,  # only need URLs, not full content
    }

    try:
        r = SESSION.post(
            "https://api.parallel.ai/v1beta/search",
            json=payload,
            headers={
                "x-api-key": PARALLEL_API_KEY,
                "parallel-beta": "search-extract-2025-10-10",
            },
            timeout=20,
        )
        r.raise_for_status()
        data = r.json()

        urls = []
        for result in data.get("results", []):
            url = result.get("url", "")
            if url and not any(skip in url for skip in SKIP_DOMAINS):
                urls.append(url)

        if DEBUG:
            print(f"    Parallel: {len(urls)} URLs")
            for u in urls[:4]:
                print(f"      {u}")

        return urls[:MAX_PAGES + 3]

    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 429:
            print("  Rate limited — waiting 15s...")
            time.sleep(15)
        return []
    except Exception as e:
        if DEBUG:
            print(f"  Parallel error: {e}")
        return []


def extract_image_from_page(page_url: str) -> str:
    """
    Fetch a product page and extract an image URL using this priority:
      1. og:image meta tag       — in <head>, never needs JS
      2. twitter:image meta tag  — same, very reliable
      3. JSON-LD schema image    — structured data in <script> tags
      4. First plausible <img>   — fallback
    """
    try:
        resp = SESSION.get(page_url, timeout=10, allow_redirects=True)
        if resp.status_code != 200:
            return ""

        soup = BeautifulSoup(resp.text, "lxml")

        # 1. og:image
        tag = soup.find("meta", property="og:image")
        if tag and is_good_image_url(tag.get("content", "")):
            return tag["content"].strip()

        # 2. twitter:image
        tag = soup.find("meta", attrs={"name": "twitter:image"})
        if tag and is_good_image_url(tag.get("content", "")):
            return tag["content"].strip()

        # 3. JSON-LD
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                ld = json.loads(script.string or "")
                for key in ("image", "thumbnailUrl"):
                    img = ld.get(key)
                    if isinstance(img, str) and is_good_image_url(img):
                        return img
                    if isinstance(img, list) and img:
                        candidate = img[0] if isinstance(img[0], str) else img[0].get("url", "")
                        if is_good_image_url(candidate):
                            return candidate
                    if isinstance(img, dict) and is_good_image_url(img.get("url", "")):
                        return img["url"]
            except Exception:
                pass

        # 4. <img> tags — pick best candidate
        for img_tag in soup.find_all("img", src=True):
            src = img_tag.get("src", "")
            if not src.startswith("http"):
                src = urljoin(page_url, src)
            if is_good_image_url(src):
                if any(kw in src.lower() for kw in ["product", "item", "catalog", "part"]):
                    return src

        return ""

    except Exception as e:
        if DEBUG:
            print(f"    Page error ({page_url[:55]}...): {e}")
        return ""


def search_image(vendor: str, part_number: str, description: str) -> str:
    """Full pipeline: search → fetch pages → return best image URL."""
    page_urls = find_product_urls(vendor, part_number, description)
    if not page_urls:
        return ""

    tried = 0
    for page_url in page_urls:
        if tried >= MAX_PAGES:
            break
        img_url = extract_image_from_page(page_url)
        tried += 1
        time.sleep(FETCH_DELAY)
        if img_url:
            if DEBUG:
                print(f"    ✓ {page_url[:60]}")
                print(f"      img: {img_url[:70]}")
            return img_url

    return ""


def main():
    print(f"Loading {INPUT_FILE}...")
    df = pd.read_excel(INPUT_FILE, dtype=str)
    df = df.fillna("")
    total = len(df)
    print(f"  {total} rows found.")

    cache = load_cache()
    print(f"  {len(cache)} cached entries loaded.\n")

    image_urls = []
    found_count = 0

    for i, row in df.iterrows():
        vendor      = str(row.get("SUPPLIER/VENDOR", "")).strip()
        part_number = str(row.get("PART NUMBER", "")).strip()
        description = str(row.get("DESCRIPTION/SIZE", "")).strip()
        cache_key   = f"{vendor}|{part_number}".upper()

        if cache_key in cache:
            url = cache[cache_key]
            image_urls.append(url)
            if url:
                found_count += 1
            label = "✓ cached" if url else "- cached miss"
            print(f"[{i+1:4d}/{total}] {label:20s} {vendor} / {part_number}")
            continue

        if not vendor and not part_number and not description:
            image_urls.append("")
            cache[cache_key] = ""
            print(f"[{i+1:4d}/{total}] {'- empty':20s}")
            continue

        url = search_image(vendor, part_number, description)
        image_urls.append(url)
        cache[cache_key] = url

        if url:
            found_count += 1
            print(f"[{i+1:4d}/{total}] {'✓ found':20s} {vendor} / {part_number}")
        else:
            print(f"[{i+1:4d}/{total}] {'✗ not found':20s} {vendor} / {part_number}")

        if (i + 1) % 25 == 0:
            save_cache(cache)
            pct = found_count / (i + 1) * 100
            print(f"  ── checkpoint: {found_count}/{i+1} found ({pct:.0f}%) ──")

        time.sleep(SEARCH_DELAY)

    save_cache(cache)
    pct = found_count / total * 100
    print(f"\n{'='*55}")
    print(f"  {found_count}/{total} images found ({pct:.0f}%)")
    print(f"{'='*55}\n")

    # ── Build Excel ───────────────────────────────────────────────────────────
    print(f"Building {OUTPUT_FILE}...")
    df.insert(0, "IMAGE", [f'=IMAGE("{u}")' if u else "" for u in image_urls])
    df.to_excel(OUTPUT_FILE, index=False)

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    ws.title = "Master Parts"

    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    even_fill = PatternFill("solid", fgColor="DCE6F1")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin
    ws.row_dimensions[1].height = 30

    for ridx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = even_fill if ridx % 2 == 0 else odd_fill
        for cell in row:
            cell.fill = fill
            cell.font = Font(name="Arial", size=10)
            cell.border = thin
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if cell.column == 1 else "left",
            )

    for col, w in {"A":18,"B":24,"C":22,"D":14,"E":50,"F":12,"G":12,"H":12,"I":12,"J":14,"K":14,"L":12,"M":30}.items():
        ws.column_dimensions[col].width = w

    for ridx in range(2, ws.max_row + 1):
        ws.row_dimensions[ridx].height = 60

    for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
        for cell in row:
            if cell.value:
                cell.number_format = "$#,##0.0000"
    for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):
        for cell in row:
            if cell.value:
                cell.number_format = "MM/DD/YYYY"

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(OUTPUT_FILE)

    print(f"Saved: {OUTPUT_FILE}")
    print(f"\nOpen in Excel 365 with internet — =IMAGE() loads automatically.")


if __name__ == "__main__":
    main()